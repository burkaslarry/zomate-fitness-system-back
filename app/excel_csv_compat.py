"""[F001][S003] / [F004][S002]
Feature: Student onboarding & payment records
Step: Excel/CSV column aliases for Fung 收錢template + system headers
Logic: Normalize gender, emergency contact, and payment row fields for dual-format import/export.
      Also convert .xlsx / .xls (first sheet only) into CSV text for the same importers.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any


def csv_cell(row: dict[str, Any], *keys: str) -> str:
    """[F001][S003] First non-empty value among alias keys (case-sensitive then stripped)."""
    for key in keys:
        if key in row and row[key] is not None:
            val = str(row[key]).strip()
            if val:
                return val
    # Case-insensitive fallback
    lower_map = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
    for key in keys:
        raw = lower_map.get(key.lower())
        if raw is not None:
            val = str(raw).strip()
            if val:
                return val
    return ""


def normalize_gender(raw: str | None) -> str | None:
    """[F001][S001] Map 男/女/M/F/male/female → male|female."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if s in ("male", "m", "男", "man", "boy"):
        return "male"
    if s in ("female", "f", "女", "woman", "girl"):
        return "female"
    return None


def gender_label_zh(gender: str | None) -> str:
    if gender == "male":
        return "男"
    if gender == "female":
        return "女"
    return ""


def parse_emergency_contact(raw: str | None) -> tuple[str | None, str | None]:
    """[F001][S001] ``Name (Relation)`` → (name, relationship)."""
    if not raw:
        return None, None
    text = str(raw).strip()
    if not text:
        return None, None
    m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", text)
    if m:
        name = m.group(1).strip() or None
        rel = m.group(2).strip() or None
        return name, rel
    return text, None


def format_emergency_contact(name: str | None, relationship: str | None) -> str:
    name = (name or "").strip()
    rel = (relationship or "").strip()
    if name and rel:
        return f"{name} ({rel})"
    return name or rel


def parse_flexible_date(raw: str | None) -> date | None:
    """[F001][S003] Accept YYYY-MM-DD, YYYY/MM/DD, M/D/YYYY, or Excel-ish mm/dd."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Bare mm/dd → assume current year (Fung sheet often shows 12/4)
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            return date(date.today().year, month, day)
        except ValueError:
            return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def parse_amount(raw: str | None) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().replace(",", "").replace("HKD", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_installment_label(raw: str | None) -> tuple[int | None, int | None]:
    """[F004][S002] ``1/3 payment`` → (1, 3)."""
    if not raw:
        return None, None
    m = re.search(r"(\d+)\s*/\s*(\d+)", str(raw))
    if not m:
        return None, None
    return int(m.group(1)), int(m.group(2))


def is_fung_student_headers(fieldnames: list[str] | None) -> bool:
    if not fieldnames:
        return False
    keys = {h.strip().lower() for h in fieldnames if h}
    return "name" in keys or "membercode" in keys or "phone number" in keys


def is_fung_payment_headers(fieldnames: list[str] | None) -> bool:
    if not fieldnames:
        return False
    keys = {h.strip() for h in fieldnames if h}
    joined = " ".join(keys)
    if "電話號碼" in keys or "收費" in keys or "付款方法" in keys:
        return True
    if "New/Renewal" in keys or "New/ Renewal" in keys:
        return True
    if "MemberCode" in keys and ("課程" in keys or "教練" in keys):
        return True
    # Header sometimes embeds date format in the name, e.g. 日期yyyy/mm/dd
    if any(k.startswith("日期") for k in keys) and ("會員" in keys or "課程" in keys):
        return True
    return "Trasaction Code" in joined or "Transaction Code" in joined


def phone_display_member_code(phone: str | None) -> str:
    """[F001][S003] Fung MemberCode style: local 8 digits with space (XXXX XXXX)."""
    if not phone:
        return ""
    digits = re.sub(r"\D", "", phone)
    if digits.startswith("852") and len(digits) >= 11:
        digits = digits[-8:]
    if len(digits) == 8:
        return f"{digits[:4]} {digits[4:]}"
    return phone


FUNG_STUDENT_HEADERS = [
    "MemberCode",
    "PT",
    "Name",
    "Gender",
    "phone number",
    "Date of Birth",
    "ID Number",
    "Emergency Contact",
    "Emergency Contact Number",
]

FUNG_PAYMENT_HEADERS = [
    "日期",
    "MemberCode",
    "會員",
    "電話號碼",
    "課程",
    "教練",
    "New/Renewal",
    "收費",
    "付款方法",
    "分期付款",
]


def _excel_cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _rows_to_csv_text(rows: list[list[Any]]) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    for row in rows:
        w.writerow([_excel_cell_to_str(c) for c in row])
    return buf.getvalue()


def _is_xlsx_bytes(raw: bytes) -> bool:
    return len(raw) >= 4 and raw[:2] == b"PK"


def _is_xls_bytes(raw: bytes) -> bool:
    # OLE Compound Document magic (classic .xls)
    return len(raw) >= 8 and raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def xlsx_first_sheet_to_csv(raw: bytes) -> str:
    """[F001][S003] Read first worksheet of an .xlsx into CSV text."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Server missing openpyxl; cannot read .xlsx.") from exc
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        if not wb.worksheets:
            raise ValueError("Excel workbook has no sheets.")
        ws = wb.worksheets[0]
        rows: list[list[Any]] = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    # Drop fully empty trailing rows
    while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        raise ValueError("Excel first sheet is empty.")
    return _rows_to_csv_text(rows)


def xls_first_sheet_to_csv(raw: bytes) -> str:
    """[F001][S003] Read first worksheet of a classic .xls into CSV text."""
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Server missing xlrd; please save as .xlsx and re-upload.") from exc
    book = xlrd.open_workbook(file_contents=raw)
    if book.nsheets < 1:
        raise ValueError("Excel workbook has no sheets.")
    sheet = book.sheet_by_index(0)
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        cells: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    cells.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    cells.append(cell.value)
            else:
                cells.append(cell.value)
        rows.append(cells)
    while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        raise ValueError("Excel first sheet is empty.")
    return _rows_to_csv_text(rows)


def upload_bytes_to_csv_text(raw: bytes, filename: str | None = None) -> str:
    """[F001][S003] CSV passthrough, or convert first sheet of .xlsx / .xls to CSV text."""
    name = (filename or "").lower().strip()
    if name.endswith((".xlsx", ".xlsm")) or (
        not name.endswith((".csv", ".txt", ".xls")) and _is_xlsx_bytes(raw)
    ):
        return xlsx_first_sheet_to_csv(raw)
    if name.endswith(".xls") or _is_xls_bytes(raw):
        # Some misnamed files are actually xlsx
        if _is_xlsx_bytes(raw):
            return xlsx_first_sheet_to_csv(raw)
        return xls_first_sheet_to_csv(raw)
    if _is_xlsx_bytes(raw):
        return xlsx_first_sheet_to_csv(raw)
    if _is_xls_bytes(raw):
        return xls_first_sheet_to_csv(raw)
    return raw.decode("utf-8-sig")
